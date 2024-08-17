from regression_algorithm import RegressionModule
from MonteCarlo_Regression.data_loader_xlsx import DataLoader
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def highlight_small_data_algorithms(excel_file):
    # Load the workbook and the active worksheet
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Define a fill pattern for highlighting (orange background)
    header_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    
    # Define bold font style
    bold_font = Font(bold=True)
    
    # Highlight the header row with orange background and bold text
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
    
    # Define a fill pattern for highlighting suitable algorithms (yellow in this case)
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Highlight the rows for suitable algorithms
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            algorithm = cell.value
            if algorithm in ['LR', 'PR', 'RR', 'Lasso', 'EN', 'BLR', 'BRR', 'KRR']:
                cell.fill = highlight_fill

    # Save the workbook
    wb.save(excel_file)


if __name__ == '__main__':
    dl = DataLoader('H-Pt(100)_Eads.xlsx')
    X, y = dl.prepare_data()

    ra = RegressionModule()
    ra.set_data(X, y)

    # Define algorithms and results container
    algorithms = [
        'LR', 'PR', 'RR', 'Lasso', 'EN', 'DTR', 'SVR', 'RFR',
        'BLR', 'PCR', 'PLSR', 'GBR', 'SGD', 'BRR', 
        'KRR', 'XGB', 'LGBM'
    ]

    results = []

    # Run each algorithm and collect results
    for algorithm in algorithms:
        print("Starting: ", algorithm)
        mse, r2, mae, rmse, mape = ra.run_regression(algorithm)
        results.append([algorithm, mse, r2, mae, rmse, mape])

    # Create a DataFrame and save to Excel
    df = pd.DataFrame(results, columns=['Algorithm', 'MSE', 'R2', 'MAE', 'RMSE', 'MAPE'])
    df.to_excel('regression_results.xlsx', index=False)

    print("Results have been saved to 'regression_results.xlsx'.")
